import openpyxl
wb = openpyxl.Workbook() #openpyxl.Workbook() W는 대문자
wb.create_sheet()
#wb.get_sheet_names()
#wb.save('test.xlsx')

#ws = openpyxl.Workbook()
sheet = wb.active
sheet['A1']='test'
wb.save('test.xlsx')
